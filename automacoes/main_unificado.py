# main_unificado.py

import sys
import csv
from faker import Faker
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def exibir_informacoes_sistema():
    print("----- INFORMAÇÕES DO SISTEMA -----")
    print(f"Plataforma: {sys.platform}")
    print(f"Python: {sys.version}")


def gerar_dados_falsos_csv(quantidade=10):
    fake = Faker("pt-BR")

    nome_csv = "dados_falsos_organizados.csv"
    nome_xlsx = "dados_falsos_organizados.xlsx"

    campos = [
        "Nome", "Email", "CPF", "RG", "Telefone",
        "Rua", "Numero", "Cidade", "Estado", "CEP",
        "Cartao", "Validade", "CVV"
    ]

    dados = []

    for _ in range(quantidade):
        endereco = fake.address().replace("\n", ", ").split(", ")

        dados.append({
            "Nome": fake.name(),
            "Email": fake.email(),
            "CPF": fake.cpf(),
            "RG": fake.rg(),
            "Telefone": fake.phone_number(),
            "Rua": endereco[0] if len(endereco) > 0 else "",
            "Numero": fake.building_number(),
            "Cidade": fake.city(),
            "Estado": fake.state_abbr(),
            "CEP": fake.postcode(),
            "Cartao": fake.credit_card_number(),
            "Validade": fake.credit_card_expire(),
            "CVV": fake.credit_card_security_code()
        })

    # CSV limpo para Excel brasileiro
    with open(nome_csv, "w", newline="", encoding="utf-8-sig") as arquivo:
        writer = csv.DictWriter(arquivo, fieldnames=campos, delimiter=";")
        writer.writeheader()
        writer.writerows(dados)

    # XLSX colorido
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"

    ws.append(campos)

    for item in dados:
        ws.append([item[campo] for campo in campos])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center")

    widths = {
        "A": 24, "B": 28, "C": 18, "D": 16, "E": 18,
        "F": 28, "G": 10, "H": 18, "I": 12, "J": 14,
        "K": 22, "L": 14, "M": 8
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    wb.save(nome_xlsx)

    return nome_csv, nome_xlsx


if __name__ == "__main__":
    exibir_informacoes_sistema()
    gerar_dados_falsos_csv(20)
    print("\nFinalizado.")